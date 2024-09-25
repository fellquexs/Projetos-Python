import os
#import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from datetime import date
from dateutil.relativedelta import relativedelta
import openpyxl
from statistics import mean

#Manipulando as datas está setado para um delta de 7 dias (usuário poderá alterar no [RANGE])--------------------------------------
data_atual = date.today()

list_datas = []

for i in range(1, 8, 1):

    aux = data_atual - relativedelta(days=i)
    aux = aux.strftime('%d-%m-%Y')
    list_datas.append(aux)
#print(f'IPDOS que serão analisados{list_datas}')

ipdoAtual = list_datas[0]
ipdo2 = list_datas[1]
ipdo3 = list_datas[2]
ipdo4 = list_datas[3]
ipdo5 = list_datas[4]
ipdo6 = list_datas[5]
IpdoComparacao = list_datas[6]

#Meu diretório-------------------------------------------------------------------------------------
diretorio = os.path.dirname(os.path.abspath(__file__))

#diretório dos IPDOs - REDE
ipdo_dirs = 'T:\Ferramentas'+ os.sep +'1_MIDDLE_2W\Whatsapp Diário\IPDO_editaveis'

atual = ipdo_dirs + os.sep + 'IPDO-' + ipdoAtual + '.xlsm' #IPDO mais RECENTE
dia2 = ipdo_dirs + os.sep + 'IPDO-' + ipdo2 + '.xlsm'
dia3 = ipdo_dirs + os.sep + 'IPDO-' + ipdo3 + '.xlsm'
dia4 = ipdo_dirs + os.sep + 'IPDO-' + ipdo4 + '.xlsm'
dia5 = ipdo_dirs + os.sep + 'IPDO-' + ipdo5 + '.xlsm'
dia6 = ipdo_dirs + os.sep + 'IPDO-' + ipdo6 + '.xlsm'
comparacao = ipdo_dirs + os.sep + 'IPDO-' + IpdoComparacao + '.xlsm' #ULTIMO IPDO

# Checking IPDO---------------------------------------------------------------------------------------

for arch in os.listdir(ipdo_dirs):
    for existe in list_datas:
        aux = arch.find(existe)
        if aux != -1 and arch.endswith('.xlsm'): #-1 Não existe , DIF DE -1 Existe
            print(arch)

#Lendo IPDO Atual-------------------------------------------------------------------------------------

wb_ipdoATUAL = load_workbook(atual)

ws_ipdo_ATUAL = wb_ipdoATUAL['IPDO']
ws_ipdo_ATUAL = wb_ipdoATUAL.active

for row in ws_ipdo_ATUAL.iter_rows(1,4000):
    for cell in row:
        if cell.value == 'Furnas':
            VolumeU_Furnas_Atual = ws_ipdo_ATUAL.cell(row=cell.row, column=20).value
            Aflu_Furnas_Atual = ws_ipdo_ATUAL.cell(row=cell.row, column=16).value
            Deflu_Furnas_Atual = ws_ipdo_ATUAL.cell(row=cell.row, column=17).value
            print(f'Furnas {ipdoAtual} \n VU - {VolumeU_Furnas_Atual}, Aflu - {Aflu_Furnas_Atual}, Deflu - {Deflu_Furnas_Atual}')

        if cell.value == 'Itumbiara':
            VolumeU_Itumbiara_Atual = ws_ipdo_ATUAL.cell(row=cell.row, column=20).value
            Aflu_Itumbiara_Atual = ws_ipdo_ATUAL.cell(row=cell.row, column=16).value
            Deflu_Itumbiara_Atual = ws_ipdo_ATUAL.cell(row=cell.row, column=17).value
            print(f'Itumbiara {ipdoAtual} \n VU - {VolumeU_Itumbiara_Atual}, Aflu - {Aflu_Itumbiara_Atual}, Deflu - {Deflu_Itumbiara_Atual}')

        if cell.value == 'Emborcação':
            VolumeU_Emborcacao_Atual = ws_ipdo_ATUAL.cell(row=cell.row, column=20).value
            Aflu_Emborcacao_Atual = ws_ipdo_ATUAL.cell(row=cell.row, column=16).value
            Deflu_Emborcacao_Atual = ws_ipdo_ATUAL.cell(row=cell.row, column=17).value
            print(f'Emborcação {ipdoAtual} \n VU - {VolumeU_Emborcacao_Atual}, Aflu - {Aflu_Emborcacao_Atual}, Deflu - {Deflu_Emborcacao_Atual}')

        if cell.value == 'Nova Ponte':
            VolumeU_NovaPonte_Atual = ws_ipdo_ATUAL.cell(row=cell.row, column=20).value
            Aflu_NovaPonte_Atual = ws_ipdo_ATUAL.cell(row=cell.row, column=16).value
            Deflu_NovaPonte_Atual = ws_ipdo_ATUAL.cell(row=cell.row, column=17).value
            print(f'Nova Ponte {ipdoAtual} \n VU - {VolumeU_NovaPonte_Atual}, Aflu - {Aflu_NovaPonte_Atual}, Deflu - {Deflu_NovaPonte_Atual}')

        if cell.value == 'Serra da Mesa':
            VolumeU_SerradaMesa_Atual = ws_ipdo_ATUAL.cell(row=cell.row, column=20).value
            Aflu_SerradaMesa_Atual = ws_ipdo_ATUAL.cell(row=cell.row, column=16).value
            Deflu_SerradaMesa_Atual = ws_ipdo_ATUAL.cell(row=cell.row, column=17).value
            print(f'Serra da Mesa {ipdoAtual} \n VU - {VolumeU_SerradaMesa_Atual}, Aflu - {Aflu_SerradaMesa_Atual}, Deflu - {Deflu_SerradaMesa_Atual}')

        if cell.value == 'G. B. Munhoz':
            VolumeU_GBMunhoz_Atual = ws_ipdo_ATUAL.cell(row=cell.row, column=20).value
            Aflu_GBMunhoz_Atual = ws_ipdo_ATUAL.cell(row=cell.row, column=16).value
            Deflu_GBMunhoz_Atual = ws_ipdo_ATUAL.cell(row=cell.row, column=17).value
            print(f'G. B. Munhoz {ipdoAtual} \n VU - {VolumeU_GBMunhoz_Atual}, Aflu - {Aflu_GBMunhoz_Atual}, Deflu - {Deflu_GBMunhoz_Atual}')

        if cell.value == 'S. Santiago':
            VolumeU_SSantiago_Atual = ws_ipdo_ATUAL.cell(row=cell.row, column=20).value
            Aflu_SSantiago_Atual = ws_ipdo_ATUAL.cell(row=cell.row, column=16).value
            Deflu_SSantiago_Atual = ws_ipdo_ATUAL.cell(row=cell.row, column=17).value
            print(f'S. Santiago {ipdoAtual} \n VU - {VolumeU_SSantiago_Atual}, Aflu - {Aflu_SSantiago_Atual}, Deflu - {Deflu_SSantiago_Atual}')

        if cell.value == 'Barra Grande':
            VolumeU_BarraGrande_Atual = ws_ipdo_ATUAL.cell(row=cell.row, column=20).value
            Aflu_BarraGrande_Atual = ws_ipdo_ATUAL.cell(row=cell.row, column=16).value
            Deflu_BarraGrande_Atual = ws_ipdo_ATUAL.cell(row=cell.row, column=17).value
            print(f'Barra Grande {ipdoAtual} \n VU - {VolumeU_BarraGrande_Atual}, Aflu - {Aflu_BarraGrande_Atual}, Deflu - {Deflu_BarraGrande_Atual}')

        if cell.value == 'Sobradinho':
            VolumeU_Sobradinho_Atual = ws_ipdo_ATUAL.cell(row=cell.row, column=20).value
            Aflu_Sobradinho_Atual = ws_ipdo_ATUAL.cell(row=cell.row, column=16).value
            Deflu_Sobradinho_Atual = ws_ipdo_ATUAL.cell(row=cell.row, column=17).value
            print(f'Sobradinho {ipdoAtual} \n VU - {VolumeU_Sobradinho_Atual}, Aflu - {Aflu_Sobradinho_Atual}, Deflu - {Deflu_Sobradinho_Atual}')

        if cell.value == 'Três Marias':
            VolumeU_TrMarias_Atual = ws_ipdo_ATUAL.cell(row=cell.row, column=20).value
            Aflu_TrMarias_Atual = ws_ipdo_ATUAL.cell(row=cell.row, column=16).value
            Deflu_TrMarias_Atual = ws_ipdo_ATUAL.cell(row=cell.row, column=17).value
            print(f'Três Marias {ipdoAtual} \n VU - {VolumeU_TrMarias_Atual}, Aflu - {Aflu_TrMarias_Atual}, Deflu - {Deflu_TrMarias_Atual}')

        if cell.value == 'Tucuruí':
            VolumeU_Tucurui_Atual = ws_ipdo_ATUAL.cell(row=cell.row, column=20).value
            Aflu_Tucurui_Atual = ws_ipdo_ATUAL.cell(row=cell.row, column=16).value
            Deflu_Tucurui_Atual = ws_ipdo_ATUAL.cell(row=cell.row, column=17).value
            print(f'Tucuruí {ipdoAtual} \n VU - {VolumeU_Tucurui_Atual}, Aflu - {Aflu_Tucurui_Atual}, Deflu - {Deflu_Tucurui_Atual}')

        if cell.value == 'Jacuí':
            VolumeU_Jacui_Atual = ws_ipdo_ATUAL.cell(row=cell.row, column=20).value
            Aflu_Jacui_Atual = ws_ipdo_ATUAL.cell(row=cell.row, column=16).value
            Deflu_Jacui_Atual = ws_ipdo_ATUAL.cell(row=cell.row, column=17).value
            print(f'Jacuí {ipdoAtual} \n VU - {VolumeU_Jacui_Atual}, Aflu - {Aflu_Jacui_Atual}, Deflu - {Deflu_Jacui_Atual}')

#Lendo IPDO Comparação--------------------------------------------------------------------------------------------------------------------------------------

wb_ipdoCOMP = load_workbook(comparacao)

ws_ipdo_COMP = wb_ipdoCOMP['IPDO']
ws_ipdo_COMP = wb_ipdoCOMP.active

for row in ws_ipdo_COMP.iter_rows(1,4000):
    for cell in row:
        if cell.value == 'Furnas':
            VolumeU_Furnas_COMP = ws_ipdo_COMP.cell(row=cell.row, column=20).value
            Aflu_Furnas_COMP = ws_ipdo_COMP.cell(row=cell.row, column=16).value
            Deflu_Furnas_COMP = ws_ipdo_COMP.cell(row=cell.row, column=17).value
            #print(f'Furnas {IpdoComparacao} \n VU - {VolumeU_Furnas_COMP}, Aflu - {Aflu_Furnas_COMP}, Deflu - {Deflu_Furnas_COMP}')

        if cell.value == 'Itumbiara':
            VolumeU_Itumbiara_COMP = ws_ipdo_COMP.cell(row=cell.row, column=20).value
            Aflu_Itumbiara_COMP = ws_ipdo_COMP.cell(row=cell.row, column=16).value
            Deflu_Itumbiara_COMP = ws_ipdo_COMP.cell(row=cell.row, column=17).value
            #print(f'Itumbiara {IpdoComparacao} \n VU - {VolumeU_Itumbiara_COMP}, Aflu - {Aflu_Itumbiara_COMP}, Deflu - {Deflu_Itumbiara_COMP}')

        if cell.value == 'Emborcação':
            VolumeU_Emborcacao_COMP = ws_ipdo_COMP.cell(row=cell.row, column=20).value
            Aflu_Emborcacao_COMP = ws_ipdo_COMP.cell(row=cell.row, column=16).value
            Deflu_Emborcacao_COMP = ws_ipdo_COMP.cell(row=cell.row, column=17).value
            #print(f'Emborcação {IpdoComparacao} \n VU - {VolumeU_Emborcacao_COMP}, Aflu - {Aflu_Emborcacao_COMP}, Deflu - {Deflu_Emborcacao_COMP}')

        if cell.value == 'Nova Ponte':
            VolumeU_NovaPonte_COMP = ws_ipdo_COMP.cell(row=cell.row, column=20).value
            Aflu_NovaPonte_COMP = ws_ipdo_COMP.cell(row=cell.row, column=16).value
            Deflu_NovaPonte_COMP = ws_ipdo_COMP.cell(row=cell.row, column=17).value
            #print(f'Nova Ponte {IpdoComparacao} \n VU - {VolumeU_NovaPonte_COMP}, Aflu - {Aflu_NovaPonte_COMP}, Deflu - {Deflu_NovaPonte_COMP}')

        if cell.value == 'Serra da Mesa':
            VolumeU_SerradaMesa_COMP = ws_ipdo_COMP.cell(row=cell.row, column=20).value
            Aflu_SerradaMesa_COMP = ws_ipdo_COMP.cell(row=cell.row, column=16).value
            Deflu_SerradaMesa_COMP = ws_ipdo_COMP.cell(row=cell.row, column=17).value
            #print(f'Serra da Mesa {IpdoComparacao} \n VU - {VolumeU_SerradaMesa_COMP}, Aflu - {Aflu_SerradaMesa_COMP}, Deflu - {Deflu_SerradaMesa_COMP}')

        if cell.value == 'G. B. Munhoz':
            VolumeU_GBMunhoz_COMP = ws_ipdo_COMP.cell(row=cell.row, column=20).value
            Aflu_GBMunhoz_COMP = ws_ipdo_COMP.cell(row=cell.row, column=16).value
            Deflu_GBMunhoz_COMP = ws_ipdo_COMP.cell(row=cell.row, column=17).value
            #print(f'G. B. Munhoz {IpdoComparacao} \n VU - {VolumeU_GBMunhoz_COMP}, Aflu - {Aflu_GBMunhoz_COMP}, Deflu - {Deflu_GBMunhoz_COMP}')

        if cell.value == 'S. Santiago':
            VolumeU_SSantiago_COMP = ws_ipdo_COMP.cell(row=cell.row, column=20).value
            Aflu_SSantiago_COMP = ws_ipdo_COMP.cell(row=cell.row, column=16).value
            Deflu_SSantiago_COMP = ws_ipdo_COMP.cell(row=cell.row, column=17).value
            #print(f'S. Santiago {IpdoComparacao} \n VU - {VolumeU_SSantiago_COMP}, Aflu - {Aflu_SSantiago_COMP}, Deflu - {Deflu_SSantiago_COMP}')

        if cell.value == 'Barra Grande':
            VolumeU_BarraGrande_COMP = ws_ipdo_COMP.cell(row=cell.row, column=20).value
            Aflu_BarraGrande_COMP = ws_ipdo_COMP.cell(row=cell.row, column=16).value
            Deflu_BarraGrande_COMP = ws_ipdo_COMP.cell(row=cell.row, column=17).value
            #print(f'Barra Grande {IpdoComparacao} \n VU - {VolumeU_BarraGrande_COMP}, Aflu - {Aflu_BarraGrande_COMP}, Deflu - {Deflu_BarraGrande_COMP}')

        if cell.value == 'Sobradinho':
            VolumeU_Sobradinho_COMP = ws_ipdo_COMP.cell(row=cell.row, column=20).value
            Aflu_Sobradinho_COMP = ws_ipdo_COMP.cell(row=cell.row, column=16).value
            Deflu_Sobradinho_COMP = ws_ipdo_COMP.cell(row=cell.row, column=17).value
            #print(f'Sobradinho {IpdoComparacao} \n VU - {VolumeU_Sobradinho_COMP}, Aflu - {Aflu_Sobradinho_COMP}, Deflu - {Deflu_Sobradinho_COMP}')

        if cell.value == 'Três Marias':
            VolumeU_TrMarias_COMP = ws_ipdo_COMP.cell(row=cell.row, column=20).value
            Aflu_TrMarias_COMP = ws_ipdo_COMP.cell(row=cell.row, column=16).value
            Deflu_TrMarias_COMP = ws_ipdo_COMP.cell(row=cell.row, column=17).value
            #print(f'Três Marias {IpdoComparacao} \n VU - {VolumeU_TrMarias_COMP}, Aflu - {Aflu_TrMarias_COMP}, Deflu - {Deflu_TrMarias_COMP}')

        if cell.value == 'Tucuruí':
            VolumeU_Tucurui_COMP = ws_ipdo_COMP.cell(row=cell.row, column=20).value
            Aflu_Tucurui_COMP = ws_ipdo_COMP.cell(row=cell.row, column=16).value
            Deflu_Tucurui_COMP = ws_ipdo_COMP.cell(row=cell.row, column=17).value
            #print(f'Tucuruí {IpdoComparacao} \n VU - {VolumeU_Tucurui_COMP}, Aflu - {Aflu_Tucurui_COMP}, Deflu - {Deflu_Tucurui_COMP}')

        if cell.value == 'Jacuí':
            VolumeU_Jacui_COMP  = ws_ipdo_COMP.cell(row=cell.row, column=20).value
            Aflu_Jacui_COMP = ws_ipdo_COMP.cell(row=cell.row, column=16).value
            Deflu_Jacui_COMP = ws_ipdo_COMP.cell(row=cell.row, column=17).value
            #print(f'Jacuí {IpdoComparacao} \n VU - {VolumeU_Jacui_COMP}, Aflu - {Aflu_Jacui_COMP}, Deflu - {Deflu_Jacui_COMP}')

#Lendo IPDO dia 2--------------------------------------------------------------------------------------------------------------------------------------

wb_ipdodia2 = load_workbook(dia2)

ws_ipdo_dia2 = wb_ipdodia2['IPDO']
ws_ipdo_dia2 = wb_ipdodia2.active

for row in ws_ipdo_dia2.iter_rows(1,4000):
    for cell in row:
        if cell.value == 'Furnas':
            VolumeU_Furnas_dia2 = ws_ipdo_dia2.cell(row=cell.row, column=20).value
            Aflu_Furnas_dia2 = ws_ipdo_dia2.cell(row=cell.row, column=16).value
            Deflu_Furnas_dia2 = ws_ipdo_dia2.cell(row=cell.row, column=17).value
            #print(f'Furnas {IpdoComparacao} \n VU - {VolumeU_Furnas_COMP}, Aflu - {Aflu_Furnas_COMP}, Deflu - {Deflu_Furnas_COMP}')
            

        if cell.value == 'Itumbiara':
            VolumeU_Itumbiara_dia2 = ws_ipdo_dia2.cell(row=cell.row, column=20).value
            Aflu_Itumbiara_dia2 = ws_ipdo_dia2.cell(row=cell.row, column=16).value
            Deflu_Itumbiara_dia2 = ws_ipdo_dia2.cell(row=cell.row, column=17).value
            #print(f'Itumbiara {IpdoComparacao} \n VU - {VolumeU_Itumbiara_COMP}, Aflu - {Aflu_Itumbiara_COMP}, Deflu - {Deflu_Itumbiara_COMP}')

        if cell.value == 'Emborcação':
            VolumeU_Emborcacao_dia2 = ws_ipdo_dia2.cell(row=cell.row, column=20).value
            Aflu_Emborcacao_dia2 = ws_ipdo_dia2.cell(row=cell.row, column=16).value
            Deflu_Emborcacao_dia2 = ws_ipdo_dia2.cell(row=cell.row, column=17).value
            #print(f'Emborcação {IpdoComparacao} \n VU - {VolumeU_Emborcacao_COMP}, Aflu - {Aflu_Emborcacao_COMP}, Deflu - {Deflu_Emborcacao_COMP}')

        if cell.value == 'Nova Ponte':
            VolumeU_NovaPonte_dia2 = ws_ipdo_dia2.cell(row=cell.row, column=20).value
            Aflu_NovaPonte_dia2 = ws_ipdo_dia2.cell(row=cell.row, column=16).value
            Deflu_NovaPonte_dia2 = ws_ipdo_dia2.cell(row=cell.row, column=17).value
            #print(f'Nova Ponte {IpdoComparacao} \n VU - {VolumeU_NovaPonte_COMP}, Aflu - {Aflu_NovaPonte_COMP}, Deflu - {Deflu_NovaPonte_COMP}')

        if cell.value == 'Serra da Mesa':
            VolumeU_SerradaMesa_dia2 = ws_ipdo_dia2.cell(row=cell.row, column=20).value
            Aflu_SerradaMesa_dia2 = ws_ipdo_dia2.cell(row=cell.row, column=16).value
            Deflu_SerradaMesa_dia2 = ws_ipdo_dia2.cell(row=cell.row, column=17).value
            #print(f'Serra da Mesa {IpdoComparacao} \n VU - {VolumeU_SerradaMesa_COMP}, Aflu - {Aflu_SerradaMesa_COMP}, Deflu - {Deflu_SerradaMesa_COMP}')

        if cell.value == 'G. B. Munhoz':
            VolumeU_GBMunhoz_dia2 = ws_ipdo_dia2.cell(row=cell.row, column=20).value
            Aflu_GBMunhoz_dia2 = ws_ipdo_dia2.cell(row=cell.row, column=16).value
            Deflu_GBMunhoz_dia2 = ws_ipdo_dia2.cell(row=cell.row, column=17).value
            #print(f'G. B. Munhoz {IpdoComparacao} \n VU - {VolumeU_GBMunhoz_COMP}, Aflu - {Aflu_GBMunhoz_COMP}, Deflu - {Deflu_GBMunhoz_COMP}')

        if cell.value == 'S. Santiago':
            VolumeU_SSantiago_dia2 = ws_ipdo_dia2.cell(row=cell.row, column=20).value
            Aflu_SSantiago_dia2 = ws_ipdo_dia2.cell(row=cell.row, column=16).value
            Deflu_SSantiago_dia2 = ws_ipdo_dia2.cell(row=cell.row, column=17).value
            #print(f'S. Santiago {IpdoComparacao} \n VU - {VolumeU_SSantiago_COMP}, Aflu - {Aflu_SSantiago_COMP}, Deflu - {Deflu_SSantiago_COMP}')

        if cell.value == 'Barra Grande':
            VolumeU_BarraGrande_dia2 = ws_ipdo_dia2.cell(row=cell.row, column=20).value
            Aflu_BarraGrande_dia2 = ws_ipdo_dia2.cell(row=cell.row, column=16).value
            Deflu_BarraGrande_dia2 = ws_ipdo_dia2.cell(row=cell.row, column=17).value
            #print(f'Barra Grande {IpdoComparacao} \n VU - {VolumeU_BarraGrande_COMP}, Aflu - {Aflu_BarraGrande_COMP}, Deflu - {Deflu_BarraGrande_COMP}')

        if cell.value == 'Sobradinho':
            VolumeU_Sobradinho_dia2 = ws_ipdo_dia2.cell(row=cell.row, column=20).value
            Aflu_Sobradinho_dia2 = ws_ipdo_dia2.cell(row=cell.row, column=16).value
            Deflu_Sobradinho_dia2 = ws_ipdo_dia2.cell(row=cell.row, column=17).value
            #print(f'Sobradinho {IpdoComparacao} \n VU - {VolumeU_Sobradinho_COMP}, Aflu - {Aflu_Sobradinho_COMP}, Deflu - {Deflu_Sobradinho_COMP}')

        if cell.value == 'Três Marias':
            VolumeU_TrMarias_dia2 = ws_ipdo_dia2.cell(row=cell.row, column=20).value
            Aflu_TrMarias_dia2 = ws_ipdo_dia2.cell(row=cell.row, column=16).value
            Deflu_TrMarias_dia2 = ws_ipdo_dia2.cell(row=cell.row, column=17).value
            #print(f'Três Marias {IpdoComparacao} \n VU - {VolumeU_TrMarias_COMP}, Aflu - {Aflu_TrMarias_COMP}, Deflu - {Deflu_TrMarias_COMP}')

        if cell.value == 'Tucuruí':
            VolumeU_Tucurui_dia2 = ws_ipdo_dia2.cell(row=cell.row, column=20).value
            Aflu_Tucurui_dia2 = ws_ipdo_dia2.cell(row=cell.row, column=16).value
            Deflu_Tucurui_dia2 = ws_ipdo_dia2.cell(row=cell.row, column=17).value
            #print(f'Tucuruí {IpdoComparacao} \n VU - {VolumeU_Tucurui_COMP}, Aflu - {Aflu_Tucurui_COMP}, Deflu - {Deflu_Tucurui_COMP}')

        if cell.value == 'Jacuí':
            VolumeU_Jacui_dia2  = ws_ipdo_dia2.cell(row=cell.row, column=20).value
            Aflu_Jacui_dia2 = ws_ipdo_dia2.cell(row=cell.row, column=16).value
            Deflu_Jacui_dia2 = ws_ipdo_dia2.cell(row=cell.row, column=17).value
            #print(f'Jacuí {IpdoComparacao} \n VU - {VolumeU_Jacui_COMP}, Aflu - {Aflu_Jacui_COMP}, Deflu - {Deflu_Jacui_COMP}')


#Lendo IPDO dia 3--------------------------------------------------------------------------------------------------------------------------------------

wb_ipdodia3 = load_workbook(dia3)

ws_ipdo_dia3 = wb_ipdodia3['IPDO']
ws_ipdo_dia3 = wb_ipdodia3.active

for row in ws_ipdo_dia3.iter_rows(1,4000):
    for cell in row:
        if cell.value == 'Furnas':
            VolumeU_Furnas_dia3 = ws_ipdo_dia3.cell(row=cell.row, column=20).value
            Aflu_Furnas_dia3 = ws_ipdo_dia3.cell(row=cell.row, column=16).value
            Deflu_Furnas_dia3 = ws_ipdo_dia3.cell(row=cell.row, column=17).value
            #print(f'Furnas {IpdoComparacao} \n VU - {VolumeU_Furnas_COMP}, Aflu - {Aflu_Furnas_COMP}, Deflu - {Deflu_Furnas_COMP}')
            

        if cell.value == 'Itumbiara':
            VolumeU_Itumbiara_dia3 = ws_ipdo_dia3.cell(row=cell.row, column=20).value
            Aflu_Itumbiara_dia3 = ws_ipdo_dia3.cell(row=cell.row, column=16).value
            Deflu_Itumbiara_dia3 = ws_ipdo_dia3.cell(row=cell.row, column=17).value
            #print(f'Itumbiara {IpdoComparacao} \n VU - {VolumeU_Itumbiara_COMP}, Aflu - {Aflu_Itumbiara_COMP}, Deflu - {Deflu_Itumbiara_COMP}')

        if cell.value == 'Emborcação':
            VolumeU_Emborcacao_dia3 = ws_ipdo_dia3.cell(row=cell.row, column=20).value
            Aflu_Emborcacao_dia3 = ws_ipdo_dia3.cell(row=cell.row, column=16).value
            Deflu_Emborcacao_dia3 = ws_ipdo_dia3.cell(row=cell.row, column=17).value
            #print(f'Emborcação {IpdoComparacao} \n VU - {VolumeU_Emborcacao_COMP}, Aflu - {Aflu_Emborcacao_COMP}, Deflu - {Deflu_Emborcacao_COMP}')

        if cell.value == 'Nova Ponte':
            VolumeU_NovaPonte_dia3 = ws_ipdo_dia3.cell(row=cell.row, column=20).value
            Aflu_NovaPonte_dia3 = ws_ipdo_dia3.cell(row=cell.row, column=16).value
            Deflu_NovaPonte_dia3 = ws_ipdo_dia3.cell(row=cell.row, column=17).value
            #print(f'Nova Ponte {IpdoComparacao} \n VU - {VolumeU_NovaPonte_COMP}, Aflu - {Aflu_NovaPonte_COMP}, Deflu - {Deflu_NovaPonte_COMP}')

        if cell.value == 'Serra da Mesa':
            VolumeU_SerradaMesa_dia3 = ws_ipdo_dia3.cell(row=cell.row, column=20).value
            Aflu_SerradaMesa_dia3 = ws_ipdo_dia3.cell(row=cell.row, column=16).value
            Deflu_SerradaMesa_dia3 = ws_ipdo_dia3.cell(row=cell.row, column=17).value
            #print(f'Serra da Mesa {IpdoComparacao} \n VU - {VolumeU_SerradaMesa_COMP}, Aflu - {Aflu_SerradaMesa_COMP}, Deflu - {Deflu_SerradaMesa_COMP}')

        if cell.value == 'G. B. Munhoz':
            VolumeU_GBMunhoz_dia3 = ws_ipdo_dia3.cell(row=cell.row, column=20).value
            Aflu_GBMunhoz_dia3 = ws_ipdo_dia3.cell(row=cell.row, column=16).value
            Deflu_GBMunhoz_dia3 = ws_ipdo_dia3.cell(row=cell.row, column=17).value
            #print(f'G. B. Munhoz {IpdoComparacao} \n VU - {VolumeU_GBMunhoz_COMP}, Aflu - {Aflu_GBMunhoz_COMP}, Deflu - {Deflu_GBMunhoz_COMP}')

        if cell.value == 'S. Santiago':
            VolumeU_SSantiago_dia3 = ws_ipdo_dia3.cell(row=cell.row, column=20).value
            Aflu_SSantiago_dia3 = ws_ipdo_dia3.cell(row=cell.row, column=16).value
            Deflu_SSantiago_dia3 = ws_ipdo_dia3.cell(row=cell.row, column=17).value
            #print(f'S. Santiago {IpdoComparacao} \n VU - {VolumeU_SSantiago_COMP}, Aflu - {Aflu_SSantiago_COMP}, Deflu - {Deflu_SSantiago_COMP}')

        if cell.value == 'Barra Grande':
            VolumeU_BarraGrande_dia3 = ws_ipdo_dia3.cell(row=cell.row, column=20).value
            Aflu_BarraGrande_dia3 = ws_ipdo_dia3.cell(row=cell.row, column=16).value
            Deflu_BarraGrande_dia3 = ws_ipdo_dia3.cell(row=cell.row, column=17).value
            #print(f'Barra Grande {IpdoComparacao} \n VU - {VolumeU_BarraGrande_COMP}, Aflu - {Aflu_BarraGrande_COMP}, Deflu - {Deflu_BarraGrande_COMP}')

        if cell.value == 'Sobradinho':
            VolumeU_Sobradinho_dia3 = ws_ipdo_dia3.cell(row=cell.row, column=20).value
            Aflu_Sobradinho_dia3 = ws_ipdo_dia3.cell(row=cell.row, column=16).value
            Deflu_Sobradinho_dia3 = ws_ipdo_dia3.cell(row=cell.row, column=17).value
            #print(f'Sobradinho {IpdoComparacao} \n VU - {VolumeU_Sobradinho_COMP}, Aflu - {Aflu_Sobradinho_COMP}, Deflu - {Deflu_Sobradinho_COMP}')

        if cell.value == 'Três Marias':
            VolumeU_TrMarias_dia3 = ws_ipdo_dia3.cell(row=cell.row, column=20).value
            Aflu_TrMarias_dia3 = ws_ipdo_dia3.cell(row=cell.row, column=16).value
            Deflu_TrMarias_dia3 = ws_ipdo_dia3.cell(row=cell.row, column=17).value
            #print(f'Três Marias {IpdoComparacao} \n VU - {VolumeU_TrMarias_COMP}, Aflu - {Aflu_TrMarias_COMP}, Deflu - {Deflu_TrMarias_COMP}')

        if cell.value == 'Tucuruí':
            VolumeU_Tucurui_dia3 = ws_ipdo_dia3.cell(row=cell.row, column=20).value
            Aflu_Tucurui_dia3 = ws_ipdo_dia3.cell(row=cell.row, column=16).value
            Deflu_Tucurui_dia3 = ws_ipdo_dia3.cell(row=cell.row, column=17).value
            #print(f'Tucuruí {IpdoComparacao} \n VU - {VolumeU_Tucurui_COMP}, Aflu - {Aflu_Tucurui_COMP}, Deflu - {Deflu_Tucurui_COMP}')

        if cell.value == 'Jacuí':
            VolumeU_Jacui_dia3  = ws_ipdo_dia3.cell(row=cell.row, column=20).value
            Aflu_Jacui_dia3 = ws_ipdo_dia3.cell(row=cell.row, column=16).value
            Deflu_Jacui_dia3 = ws_ipdo_dia3.cell(row=cell.row, column=17).value
            #print(f'Jacuí {IpdoComparacao} \n VU - {VolumeU_Jacui_COMP}, Aflu - {Aflu_Jacui_COMP}, Deflu - {Deflu_Jacui_COMP}')



#Lendo IPDO dia 4--------------------------------------------------------------------------------------------------------------------------------------

wb_ipdodia4 = load_workbook(dia4)

ws_ipdo_dia4 = wb_ipdodia4['IPDO']
ws_ipdo_dia4 = wb_ipdodia4.active

for row in ws_ipdo_dia4.iter_rows(1,4000):
    for cell in row:
        if cell.value == 'Furnas':
            VolumeU_Furnas_dia4 = ws_ipdo_dia4.cell(row=cell.row, column=20).value
            Aflu_Furnas_dia4 = ws_ipdo_dia4.cell(row=cell.row, column=16).value
            Deflu_Furnas_dia4 = ws_ipdo_dia4.cell(row=cell.row, column=17).value
            #print(f'Furnas {IpdoComparacao} \n VU - {VolumeU_Furnas_COMP}, Aflu - {Aflu_Furnas_COMP}, Deflu - {Deflu_Furnas_COMP}')
            

        if cell.value == 'Itumbiara':
            VolumeU_Itumbiara_dia4 = ws_ipdo_dia4.cell(row=cell.row, column=20).value
            Aflu_Itumbiara_dia4 = ws_ipdo_dia4.cell(row=cell.row, column=16).value
            Deflu_Itumbiara_dia4 = ws_ipdo_dia4.cell(row=cell.row, column=17).value
            #print(f'Itumbiara {IpdoComparacao} \n VU - {VolumeU_Itumbiara_COMP}, Aflu - {Aflu_Itumbiara_COMP}, Deflu - {Deflu_Itumbiara_COMP}')

        if cell.value == 'Emborcação':
            VolumeU_Emborcacao_dia4 = ws_ipdo_dia4.cell(row=cell.row, column=20).value
            Aflu_Emborcacao_dia4 = ws_ipdo_dia4.cell(row=cell.row, column=16).value
            Deflu_Emborcacao_dia4 = ws_ipdo_dia4.cell(row=cell.row, column=17).value
            #print(f'Emborcação {IpdoComparacao} \n VU - {VolumeU_Emborcacao_COMP}, Aflu - {Aflu_Emborcacao_COMP}, Deflu - {Deflu_Emborcacao_COMP}')

        if cell.value == 'Nova Ponte':
            VolumeU_NovaPonte_dia4 = ws_ipdo_dia4.cell(row=cell.row, column=20).value
            Aflu_NovaPonte_dia4 = ws_ipdo_dia4.cell(row=cell.row, column=16).value
            Deflu_NovaPonte_dia4 = ws_ipdo_dia4.cell(row=cell.row, column=17).value
            #print(f'Nova Ponte {IpdoComparacao} \n VU - {VolumeU_NovaPonte_COMP}, Aflu - {Aflu_NovaPonte_COMP}, Deflu - {Deflu_NovaPonte_COMP}')

        if cell.value == 'Serra da Mesa':
            VolumeU_SerradaMesa_dia4 = ws_ipdo_dia4.cell(row=cell.row, column=20).value
            Aflu_SerradaMesa_dia4 = ws_ipdo_dia4.cell(row=cell.row, column=16).value
            Deflu_SerradaMesa_dia4 = ws_ipdo_dia4.cell(row=cell.row, column=17).value
            #print(f'Serra da Mesa {IpdoComparacao} \n VU - {VolumeU_SerradaMesa_COMP}, Aflu - {Aflu_SerradaMesa_COMP}, Deflu - {Deflu_SerradaMesa_COMP}')

        if cell.value == 'G. B. Munhoz':
            VolumeU_GBMunhoz_dia4 = ws_ipdo_dia4.cell(row=cell.row, column=20).value
            Aflu_GBMunhoz_dia4 = ws_ipdo_dia4.cell(row=cell.row, column=16).value
            Deflu_GBMunhoz_dia4 = ws_ipdo_dia4.cell(row=cell.row, column=17).value
            #print(f'G. B. Munhoz {IpdoComparacao} \n VU - {VolumeU_GBMunhoz_COMP}, Aflu - {Aflu_GBMunhoz_COMP}, Deflu - {Deflu_GBMunhoz_COMP}')

        if cell.value == 'S. Santiago':
            VolumeU_SSantiago_dia4 = ws_ipdo_dia4.cell(row=cell.row, column=20).value
            Aflu_SSantiago_dia4 = ws_ipdo_dia4.cell(row=cell.row, column=16).value
            Deflu_SSantiago_dia4 = ws_ipdo_dia4.cell(row=cell.row, column=17).value
            #print(f'S. Santiago {IpdoComparacao} \n VU - {VolumeU_SSantiago_COMP}, Aflu - {Aflu_SSantiago_COMP}, Deflu - {Deflu_SSantiago_COMP}')

        if cell.value == 'Barra Grande':
            VolumeU_BarraGrande_dia4 = ws_ipdo_dia4.cell(row=cell.row, column=20).value
            Aflu_BarraGrande_dia4 = ws_ipdo_dia4.cell(row=cell.row, column=16).value
            Deflu_BarraGrande_dia4 = ws_ipdo_dia4.cell(row=cell.row, column=17).value
            #print(f'Barra Grande {IpdoComparacao} \n VU - {VolumeU_BarraGrande_COMP}, Aflu - {Aflu_BarraGrande_COMP}, Deflu - {Deflu_BarraGrande_COMP}')

        if cell.value == 'Sobradinho':
            VolumeU_Sobradinho_dia4 = ws_ipdo_dia4.cell(row=cell.row, column=20).value
            Aflu_Sobradinho_dia4 = ws_ipdo_dia4.cell(row=cell.row, column=16).value
            Deflu_Sobradinho_dia4 = ws_ipdo_dia4.cell(row=cell.row, column=17).value
            #print(f'Sobradinho {IpdoComparacao} \n VU - {VolumeU_Sobradinho_COMP}, Aflu - {Aflu_Sobradinho_COMP}, Deflu - {Deflu_Sobradinho_COMP}')

        if cell.value == 'Três Marias':
            VolumeU_TrMarias_dia4 = ws_ipdo_dia4.cell(row=cell.row, column=20).value
            Aflu_TrMarias_dia4 = ws_ipdo_dia4.cell(row=cell.row, column=16).value
            Deflu_TrMarias_dia4 = ws_ipdo_dia4.cell(row=cell.row, column=17).value
            #print(f'Três Marias {IpdoComparacao} \n VU - {VolumeU_TrMarias_COMP}, Aflu - {Aflu_TrMarias_COMP}, Deflu - {Deflu_TrMarias_COMP}')

        if cell.value == 'Tucuruí':
            VolumeU_Tucurui_dia4 = ws_ipdo_dia4.cell(row=cell.row, column=20).value
            Aflu_Tucurui_dia4 = ws_ipdo_dia4.cell(row=cell.row, column=16).value
            Deflu_Tucurui_dia4 = ws_ipdo_dia4.cell(row=cell.row, column=17).value
            #print(f'Tucuruí {IpdoComparacao} \n VU - {VolumeU_Tucurui_COMP}, Aflu - {Aflu_Tucurui_COMP}, Deflu - {Deflu_Tucurui_COMP}')

        if cell.value == 'Jacuí':
            VolumeU_Jacui_dia4  = ws_ipdo_dia4.cell(row=cell.row, column=20).value
            Aflu_Jacui_dia4 = ws_ipdo_dia4.cell(row=cell.row, column=16).value
            Deflu_Jacui_dia4 = ws_ipdo_dia4.cell(row=cell.row, column=17).value
            #print(f'Jacuí {IpdoComparacao} \n VU - {VolumeU_Jacui_COMP}, Aflu - {Aflu_Jacui_COMP}, Deflu - {Deflu_Jacui_COMP}')

#Lendo IPDO dia 5--------------------------------------------------------------------------------------------------------------------------------------

wb_ipdodia5 = load_workbook(dia5)

ws_ipdo_dia5 = wb_ipdodia5['IPDO']
ws_ipdo_dia5 = wb_ipdodia5.active

for row in ws_ipdo_dia5.iter_rows(1,4000):
    for cell in row:
        if cell.value == 'Furnas':
            VolumeU_Furnas_dia5 = ws_ipdo_dia5.cell(row=cell.row, column=20).value
            Aflu_Furnas_dia5 = ws_ipdo_dia5.cell(row=cell.row, column=16).value
            Deflu_Furnas_dia5 = ws_ipdo_dia5.cell(row=cell.row, column=17).value
            #print(f'Furnas {IpdoComparacao} \n VU - {VolumeU_Furnas_COMP}, Aflu - {Aflu_Furnas_COMP}, Deflu - {Deflu_Furnas_COMP}')
            

        if cell.value == 'Itumbiara':
            VolumeU_Itumbiara_dia5 = ws_ipdo_dia5.cell(row=cell.row, column=20).value
            Aflu_Itumbiara_dia5 = ws_ipdo_dia5.cell(row=cell.row, column=16).value
            Deflu_Itumbiara_dia5 = ws_ipdo_dia5.cell(row=cell.row, column=17).value
            #print(f'Itumbiara {IpdoComparacao} \n VU - {VolumeU_Itumbiara_COMP}, Aflu - {Aflu_Itumbiara_COMP}, Deflu - {Deflu_Itumbiara_COMP}')

        if cell.value == 'Emborcação':
            VolumeU_Emborcacao_dia5 = ws_ipdo_dia5.cell(row=cell.row, column=20).value
            Aflu_Emborcacao_dia5 = ws_ipdo_dia5.cell(row=cell.row, column=16).value
            Deflu_Emborcacao_dia5 = ws_ipdo_dia5.cell(row=cell.row, column=17).value
            #print(f'Emborcação {IpdoComparacao} \n VU - {VolumeU_Emborcacao_COMP}, Aflu - {Aflu_Emborcacao_COMP}, Deflu - {Deflu_Emborcacao_COMP}')

        if cell.value == 'Nova Ponte':
            VolumeU_NovaPonte_dia5 = ws_ipdo_dia5.cell(row=cell.row, column=20).value
            Aflu_NovaPonte_dia5 = ws_ipdo_dia5.cell(row=cell.row, column=16).value
            Deflu_NovaPonte_dia5 = ws_ipdo_dia5.cell(row=cell.row, column=17).value
            #print(f'Nova Ponte {IpdoComparacao} \n VU - {VolumeU_NovaPonte_COMP}, Aflu - {Aflu_NovaPonte_COMP}, Deflu - {Deflu_NovaPonte_COMP}')

        if cell.value == 'Serra da Mesa':
            VolumeU_SerradaMesa_dia5 = ws_ipdo_dia5.cell(row=cell.row, column=20).value
            Aflu_SerradaMesa_dia5 = ws_ipdo_dia5.cell(row=cell.row, column=16).value
            Deflu_SerradaMesa_dia5 = ws_ipdo_dia5.cell(row=cell.row, column=17).value
            #print(f'Serra da Mesa {IpdoComparacao} \n VU - {VolumeU_SerradaMesa_COMP}, Aflu - {Aflu_SerradaMesa_COMP}, Deflu - {Deflu_SerradaMesa_COMP}')

        if cell.value == 'G. B. Munhoz':
            VolumeU_GBMunhoz_dia5 = ws_ipdo_dia5.cell(row=cell.row, column=20).value
            Aflu_GBMunhoz_dia5 = ws_ipdo_dia5.cell(row=cell.row, column=16).value
            Deflu_GBMunhoz_dia5 = ws_ipdo_dia5.cell(row=cell.row, column=17).value
            #print(f'G. B. Munhoz {IpdoComparacao} \n VU - {VolumeU_GBMunhoz_COMP}, Aflu - {Aflu_GBMunhoz_COMP}, Deflu - {Deflu_GBMunhoz_COMP}')

        if cell.value == 'S. Santiago':
            VolumeU_SSantiago_dia5 = ws_ipdo_dia5.cell(row=cell.row, column=20).value
            Aflu_SSantiago_dia5 = ws_ipdo_dia5.cell(row=cell.row, column=16).value
            Deflu_SSantiago_dia5 = ws_ipdo_dia5.cell(row=cell.row, column=17).value
            #print(f'S. Santiago {IpdoComparacao} \n VU - {VolumeU_SSantiago_COMP}, Aflu - {Aflu_SSantiago_COMP}, Deflu - {Deflu_SSantiago_COMP}')

        if cell.value == 'Barra Grande':
            VolumeU_BarraGrande_dia5 = ws_ipdo_dia5.cell(row=cell.row, column=20).value
            Aflu_BarraGrande_dia5 = ws_ipdo_dia5.cell(row=cell.row, column=16).value
            Deflu_BarraGrande_dia5 = ws_ipdo_dia5.cell(row=cell.row, column=17).value
            #print(f'Barra Grande {IpdoComparacao} \n VU - {VolumeU_BarraGrande_COMP}, Aflu - {Aflu_BarraGrande_COMP}, Deflu - {Deflu_BarraGrande_COMP}')

        if cell.value == 'Sobradinho':
            VolumeU_Sobradinho_dia5 = ws_ipdo_dia5.cell(row=cell.row, column=20).value
            Aflu_Sobradinho_dia5 = ws_ipdo_dia5.cell(row=cell.row, column=16).value
            Deflu_Sobradinho_dia5 = ws_ipdo_dia5.cell(row=cell.row, column=17).value
            #print(f'Sobradinho {IpdoComparacao} \n VU - {VolumeU_Sobradinho_COMP}, Aflu - {Aflu_Sobradinho_COMP}, Deflu - {Deflu_Sobradinho_COMP}')

        if cell.value == 'Três Marias':
            VolumeU_TrMarias_dia5 = ws_ipdo_dia5.cell(row=cell.row, column=20).value
            Aflu_TrMarias_dia5 = ws_ipdo_dia5.cell(row=cell.row, column=16).value
            Deflu_TrMarias_dia5 = ws_ipdo_dia5.cell(row=cell.row, column=17).value
            #print(f'Três Marias {IpdoComparacao} \n VU - {VolumeU_TrMarias_COMP}, Aflu - {Aflu_TrMarias_COMP}, Deflu - {Deflu_TrMarias_COMP}')

        if cell.value == 'Tucuruí':
            VolumeU_Tucurui_dia5 = ws_ipdo_dia5.cell(row=cell.row, column=20).value
            Aflu_Tucurui_dia5 = ws_ipdo_dia5.cell(row=cell.row, column=16).value
            Deflu_Tucurui_dia5 = ws_ipdo_dia5.cell(row=cell.row, column=17).value
            #print(f'Tucuruí {IpdoComparacao} \n VU - {VolumeU_Tucurui_COMP}, Aflu - {Aflu_Tucurui_COMP}, Deflu - {Deflu_Tucurui_COMP}')

        if cell.value == 'Jacuí':
            VolumeU_Jacui_dia5  = ws_ipdo_dia5.cell(row=cell.row, column=20).value
            Aflu_Jacui_dia5 = ws_ipdo_dia5.cell(row=cell.row, column=16).value
            Deflu_Jacui_dia5 = ws_ipdo_dia5.cell(row=cell.row, column=17).value
            #print(f'Jacuí {IpdoComparacao} \n VU - {VolumeU_Jacui_COMP}, Aflu - {Aflu_Jacui_COMP}, Deflu - {Deflu_Jacui_COMP}')

#Lendo IPDO dia 6--------------------------------------------------------------------------------------------------------------------------------------

wb_ipdodia6 = load_workbook(dia6)

ws_ipdo_dia6 = wb_ipdodia6['IPDO']
ws_ipdo_dia6 = wb_ipdodia6.active

for row in ws_ipdo_dia6.iter_rows(1,4000):
    for cell in row:
        if cell.value == 'Furnas':
            VolumeU_Furnas_dia6 = ws_ipdo_dia6.cell(row=cell.row, column=20).value
            Aflu_Furnas_dia6 = ws_ipdo_dia6.cell(row=cell.row, column=16).value
            Deflu_Furnas_dia6 = ws_ipdo_dia6.cell(row=cell.row, column=17).value
            #print(f'Furnas {IpdoComparacao} \n VU - {VolumeU_Furnas_COMP}, Aflu - {Aflu_Furnas_COMP}, Deflu - {Deflu_Furnas_COMP}')
            

        if cell.value == 'Itumbiara':
            VolumeU_Itumbiara_dia6 = ws_ipdo_dia6.cell(row=cell.row, column=20).value
            Aflu_Itumbiara_dia6 = ws_ipdo_dia6.cell(row=cell.row, column=16).value
            Deflu_Itumbiara_dia6 = ws_ipdo_dia6.cell(row=cell.row, column=17).value
            #print(f'Itumbiara {IpdoComparacao} \n VU - {VolumeU_Itumbiara_COMP}, Aflu - {Aflu_Itumbiara_COMP}, Deflu - {Deflu_Itumbiara_COMP}')

        if cell.value == 'Emborcação':
            VolumeU_Emborcacao_dia6 = ws_ipdo_dia6.cell(row=cell.row, column=20).value
            Aflu_Emborcacao_dia6 = ws_ipdo_dia6.cell(row=cell.row, column=16).value
            Deflu_Emborcacao_dia6 = ws_ipdo_dia6.cell(row=cell.row, column=17).value
            #print(f'Emborcação {IpdoComparacao} \n VU - {VolumeU_Emborcacao_COMP}, Aflu - {Aflu_Emborcacao_COMP}, Deflu - {Deflu_Emborcacao_COMP}')

        if cell.value == 'Nova Ponte':
            VolumeU_NovaPonte_dia6 = ws_ipdo_dia6.cell(row=cell.row, column=20).value
            Aflu_NovaPonte_dia6 = ws_ipdo_dia6.cell(row=cell.row, column=16).value
            Deflu_NovaPonte_dia6 = ws_ipdo_dia6.cell(row=cell.row, column=17).value
            #print(f'Nova Ponte {IpdoComparacao} \n VU - {VolumeU_NovaPonte_COMP}, Aflu - {Aflu_NovaPonte_COMP}, Deflu - {Deflu_NovaPonte_COMP}')

        if cell.value == 'Serra da Mesa':
            VolumeU_SerradaMesa_dia6 = ws_ipdo_dia6.cell(row=cell.row, column=20).value
            Aflu_SerradaMesa_dia6 = ws_ipdo_dia6.cell(row=cell.row, column=16).value
            Deflu_SerradaMesa_dia6 = ws_ipdo_dia6.cell(row=cell.row, column=17).value
            #print(f'Serra da Mesa {IpdoComparacao} \n VU - {VolumeU_SerradaMesa_COMP}, Aflu - {Aflu_SerradaMesa_COMP}, Deflu - {Deflu_SerradaMesa_COMP}')

        if cell.value == 'G. B. Munhoz':
            VolumeU_GBMunhoz_dia6 = ws_ipdo_dia6.cell(row=cell.row, column=20).value
            Aflu_GBMunhoz_dia6 = ws_ipdo_dia6.cell(row=cell.row, column=16).value
            Deflu_GBMunhoz_dia6 = ws_ipdo_dia6.cell(row=cell.row, column=17).value
            #print(f'G. B. Munhoz {IpdoComparacao} \n VU - {VolumeU_GBMunhoz_COMP}, Aflu - {Aflu_GBMunhoz_COMP}, Deflu - {Deflu_GBMunhoz_COMP}')

        if cell.value == 'S. Santiago':
            VolumeU_SSantiago_dia6 = ws_ipdo_dia6.cell(row=cell.row, column=20).value
            Aflu_SSantiago_dia6 = ws_ipdo_dia6.cell(row=cell.row, column=16).value
            Deflu_SSantiago_dia6 = ws_ipdo_dia6.cell(row=cell.row, column=17).value
            #print(f'S. Santiago {IpdoComparacao} \n VU - {VolumeU_SSantiago_COMP}, Aflu - {Aflu_SSantiago_COMP}, Deflu - {Deflu_SSantiago_COMP}')

        if cell.value == 'Barra Grande':
            VolumeU_BarraGrande_dia6 = ws_ipdo_dia6.cell(row=cell.row, column=20).value
            Aflu_BarraGrande_dia6 = ws_ipdo_dia6.cell(row=cell.row, column=16).value
            Deflu_BarraGrande_dia6 = ws_ipdo_dia6.cell(row=cell.row, column=17).value
            #print(f'Barra Grande {IpdoComparacao} \n VU - {VolumeU_BarraGrande_COMP}, Aflu - {Aflu_BarraGrande_COMP}, Deflu - {Deflu_BarraGrande_COMP}')

        if cell.value == 'Sobradinho':
            VolumeU_Sobradinho_dia6 = ws_ipdo_dia6.cell(row=cell.row, column=20).value
            Aflu_Sobradinho_dia6 = ws_ipdo_dia6.cell(row=cell.row, column=16).value
            Deflu_Sobradinho_dia6 = ws_ipdo_dia6.cell(row=cell.row, column=17).value
            #print(f'Sobradinho {IpdoComparacao} \n VU - {VolumeU_Sobradinho_COMP}, Aflu - {Aflu_Sobradinho_COMP}, Deflu - {Deflu_Sobradinho_COMP}')

        if cell.value == 'Três Marias':
            VolumeU_TrMarias_dia6 = ws_ipdo_dia6.cell(row=cell.row, column=20).value
            Aflu_TrMarias_dia6 = ws_ipdo_dia6.cell(row=cell.row, column=16).value
            Deflu_TrMarias_dia6 = ws_ipdo_dia6.cell(row=cell.row, column=17).value
            #print(f'Três Marias {IpdoComparacao} \n VU - {VolumeU_TrMarias_COMP}, Aflu - {Aflu_TrMarias_COMP}, Deflu - {Deflu_TrMarias_COMP}')

        if cell.value == 'Tucuruí':
            VolumeU_Tucurui_dia6 = ws_ipdo_dia6.cell(row=cell.row, column=20).value
            Aflu_Tucurui_dia6 = ws_ipdo_dia6.cell(row=cell.row, column=16).value
            Deflu_Tucurui_dia6 = ws_ipdo_dia6.cell(row=cell.row, column=17).value
            #print(f'Tucuruí {IpdoComparacao} \n VU - {VolumeU_Tucurui_COMP}, Aflu - {Aflu_Tucurui_COMP}, Deflu - {Deflu_Tucurui_COMP}')

        if cell.value == 'Jacuí':
            VolumeU_Jacui_dia6  = ws_ipdo_dia6.cell(row=cell.row, column=20).value
            Aflu_Jacui_dia6 = ws_ipdo_dia6.cell(row=cell.row, column=16).value
            Deflu_Jacui_dia6 = ws_ipdo_dia6.cell(row=cell.row, column=17).value
            #print(f'Jacuí {IpdoComparacao} \n VU - {VolumeU_Jacui_COMP}, Aflu - {Aflu_Jacui_COMP}, Deflu - {Deflu_Jacui_COMP}')

Mdeflu_NovaPonte = format(mean([Deflu_NovaPonte_Atual, Deflu_NovaPonte_COMP, Deflu_NovaPonte_dia2, Deflu_NovaPonte_dia3, Deflu_NovaPonte_dia4, Deflu_NovaPonte_dia5, Deflu_NovaPonte_dia6]),".2f")
Maflu_NovaPonte = format(mean([Aflu_NovaPonte_Atual, Aflu_NovaPonte_COMP, Aflu_NovaPonte_dia2, Aflu_NovaPonte_dia3, Aflu_NovaPonte_dia4, Aflu_NovaPonte_dia5, Aflu_NovaPonte_dia6]),".2f")

Maflu_Emborcacao = format(mean([Aflu_Emborcacao_Atual, Aflu_Emborcacao_COMP, Aflu_Emborcacao_dia2, Aflu_Emborcacao_dia3, Aflu_Emborcacao_dia4, Aflu_Emborcacao_dia5, Aflu_Emborcacao_dia6]),".2f")
Mdeflu_Emborcacao = format(mean([Deflu_Emborcacao_Atual, Deflu_Emborcacao_COMP, Deflu_Emborcacao_dia2, Deflu_Emborcacao_dia3, Deflu_Emborcacao_dia4, Deflu_Emborcacao_dia5, Deflu_Emborcacao_dia6]),".2f")

Maflu_Itumbiara = format(mean([Aflu_Itumbiara_Atual, Aflu_Itumbiara_COMP, Aflu_Itumbiara_dia2, Aflu_Itumbiara_dia3, Aflu_Itumbiara_dia4, Aflu_Itumbiara_dia5, Aflu_Itumbiara_dia6]),".2f")
Mdeflu_Itumbiara = format(mean([Deflu_Itumbiara_Atual, Deflu_Itumbiara_COMP, Deflu_Itumbiara_dia2, Deflu_Itumbiara_dia3, Deflu_Itumbiara_dia4, Deflu_Itumbiara_dia5, Deflu_Itumbiara_dia6]),".2f")

Maflu_Furnas = format(mean([Aflu_Furnas_Atual, Aflu_Furnas_COMP, Aflu_Furnas_dia2, Aflu_Furnas_dia3, Aflu_Furnas_dia4, Aflu_Furnas_dia5, Aflu_Furnas_dia6]),".2f")
Mdeflu_Furnas = format(mean([Deflu_Furnas_Atual, Deflu_Furnas_COMP, Deflu_Furnas_dia2, Deflu_Furnas_dia3, Deflu_Furnas_dia4, Deflu_Furnas_dia5, Deflu_Furnas_dia6]),".2f")

Maflu_GBMunhoz = format(mean([Aflu_GBMunhoz_Atual, Aflu_GBMunhoz_COMP, Aflu_GBMunhoz_dia2, Aflu_GBMunhoz_dia3, Aflu_GBMunhoz_dia4, Aflu_GBMunhoz_dia5, Aflu_GBMunhoz_dia6]),".2f")
Mdeflu_GBMunhoz = format(mean([Deflu_GBMunhoz_Atual, Deflu_GBMunhoz_COMP, Deflu_GBMunhoz_dia2, Deflu_GBMunhoz_dia3, Deflu_GBMunhoz_dia4, Deflu_GBMunhoz_dia5, Deflu_GBMunhoz_dia6]),".2f")

Maflu_SSantiago = format(mean([Aflu_SSantiago_Atual, Aflu_SSantiago_COMP, Aflu_SSantiago_dia2, Aflu_SSantiago_dia3, Aflu_SSantiago_dia4, Aflu_SSantiago_dia5, Aflu_SSantiago_dia6]),".2f")
Mdeflu_SSantiago = format(mean([Deflu_SSantiago_Atual, Deflu_SSantiago_COMP, Deflu_SSantiago_dia2, Deflu_SSantiago_dia3, Deflu_SSantiago_dia4, Deflu_SSantiago_dia5, Deflu_SSantiago_dia6]),".2f")

Maflu_BarraGrande = format(mean([Aflu_BarraGrande_Atual, Aflu_BarraGrande_COMP, Aflu_BarraGrande_dia2, Aflu_BarraGrande_dia3, Aflu_BarraGrande_dia4, Aflu_BarraGrande_dia5, Aflu_BarraGrande_dia6]),".2f")
Mdeflu_BarraGrande = format(mean([Deflu_BarraGrande_Atual, Deflu_BarraGrande_COMP, Deflu_BarraGrande_dia2, Deflu_BarraGrande_dia3, Deflu_BarraGrande_dia4, Deflu_BarraGrande_dia5, Deflu_BarraGrande_dia6]),".2f")

Maflu_Jacui = format(mean([Aflu_Jacui_Atual, Aflu_Jacui_COMP, Aflu_Jacui_dia2, Aflu_Jacui_dia3, Aflu_Jacui_dia4, Aflu_Jacui_dia5, Aflu_Jacui_dia6]),".2f")
Mdeflu_Jacui = format(mean([Deflu_Jacui_Atual, Deflu_Jacui_COMP, Deflu_Jacui_dia2, Deflu_Jacui_dia3, Deflu_Jacui_dia4, Deflu_Jacui_dia5, Deflu_Jacui_dia6]),".2f")

Maflu_SerradaMesa = format(mean([Aflu_SerradaMesa_Atual, Aflu_SerradaMesa_COMP, Aflu_SerradaMesa_dia2, Aflu_SerradaMesa_dia3, Aflu_SerradaMesa_dia4, Aflu_SerradaMesa_dia5, Aflu_SerradaMesa_dia6]),".2f")
Mdeflu_SerradaMesa = format(mean([Deflu_SerradaMesa_Atual, Deflu_SerradaMesa_COMP, Deflu_SerradaMesa_dia2, Deflu_SerradaMesa_dia3, Deflu_SerradaMesa_dia4, Deflu_SerradaMesa_dia5, Deflu_SerradaMesa_dia6]),".2f")

Maflu_Tucurui = format(mean([Aflu_Tucurui_Atual, Aflu_Tucurui_COMP, Aflu_Tucurui_dia2, Aflu_Tucurui_dia3, Aflu_Tucurui_dia4, Aflu_Tucurui_dia5, Aflu_Tucurui_dia6]),".2f")
Mdeflu_Tucurui = format(mean([Deflu_Tucurui_Atual, Deflu_Tucurui_COMP, Deflu_Tucurui_dia2, Deflu_Tucurui_dia3, Deflu_Tucurui_dia4, Deflu_Tucurui_dia5, Deflu_Tucurui_dia6]),".2f")

Maflu_TrMarias = format(mean([Aflu_TrMarias_Atual, Aflu_TrMarias_COMP, Aflu_TrMarias_dia2, Aflu_TrMarias_dia3, Aflu_TrMarias_dia4, Aflu_TrMarias_dia5, Aflu_TrMarias_dia6]),".2f")
Mdeflu_TrMarias = format(mean([Deflu_TrMarias_Atual, Deflu_TrMarias_COMP, Deflu_TrMarias_dia2, Deflu_TrMarias_dia3, Deflu_TrMarias_dia4, Deflu_TrMarias_dia5, Deflu_TrMarias_dia6]),".2f")

Maflu_Sobradinho = format(mean([Aflu_Sobradinho_Atual, Aflu_Sobradinho_COMP, Aflu_Sobradinho_dia2, Aflu_Sobradinho_dia3, Aflu_Sobradinho_dia4, Aflu_Sobradinho_dia5, Aflu_Sobradinho_dia6]),".2f")
Mdeflu_Sobradinho = format(mean([Deflu_Sobradinho_Atual, Deflu_Sobradinho_COMP, Deflu_Sobradinho_dia2, Deflu_Sobradinho_dia3, Deflu_Sobradinho_dia4, Deflu_Sobradinho_dia5, Deflu_Sobradinho_dia6]),".2f")

print(Aflu_Furnas_Atual, '-', ipdoAtual)
print(Aflu_Furnas_dia2, '-', ipdo2)
print(Aflu_Furnas_dia3, '-', ipdo3)
print(Aflu_Furnas_dia4, '-', ipdo4)
print(Aflu_Furnas_dia5, '-', ipdo5)
print(Aflu_Furnas_dia6, '-', ipdo6)
print(Aflu_Furnas_COMP, '-', IpdoComparacao)
#Inserido dados na plan2 no excel FollowSIN--------------------------------------------------------------------------------------------


followSIN = load_workbook(diretorio + os.sep + 'FollowSIN.xlsx')
page_dados = followSIN['plan2']
page_dados = followSIN.active

page_dados.append(['Bacias','Afluência m³/s','Defluência m³/s','VU','Média Afluência m³/s','Média Defluência m³/s','Delta VU %'])
page_dados.append(['Nova Ponte',Aflu_NovaPonte_Atual,Deflu_NovaPonte_Atual,VolumeU_NovaPonte_Atual, Maflu_NovaPonte, Mdeflu_NovaPonte, VolumeU_NovaPonte_Atual - VolumeU_NovaPonte_COMP])
page_dados.append(['Emborcação',Aflu_Emborcacao_Atual,Deflu_Emborcacao_Atual,VolumeU_Emborcacao_Atual,Maflu_Emborcacao, Mdeflu_Emborcacao , VolumeU_Emborcacao_Atual - VolumeU_Emborcacao_COMP])
page_dados.append(['Itumbiara',Aflu_Itumbiara_Atual,Deflu_Itumbiara_Atual,VolumeU_Itumbiara_Atual, Maflu_Itumbiara , Mdeflu_Itumbiara, VolumeU_Itumbiara_Atual - VolumeU_Itumbiara_COMP])
page_dados.append(['Furnas',Aflu_Furnas_Atual,Deflu_Furnas_Atual, VolumeU_Furnas_Atual, Maflu_Furnas, Mdeflu_Furnas , VolumeU_Furnas_Atual - VolumeU_Furnas_COMP])
page_dados.append(['G. B. Munhoz',Aflu_GBMunhoz_Atual,Deflu_GBMunhoz_Atual,VolumeU_GBMunhoz_Atual, Maflu_GBMunhoz , Mdeflu_GBMunhoz, VolumeU_GBMunhoz_Atual - VolumeU_GBMunhoz_COMP])
page_dados.append(['S. Santiago',Aflu_SSantiago_Atual,Deflu_SSantiago_Atual, VolumeU_SSantiago_Atual, Maflu_SSantiago, Mdeflu_SSantiago, VolumeU_SSantiago_Atual - VolumeU_SSantiago_COMP])
page_dados.append(['Barra Grande',Aflu_BarraGrande_Atual,Deflu_BarraGrande_Atual,VolumeU_BarraGrande_Atual, Maflu_BarraGrande, Mdeflu_BarraGrande, VolumeU_BarraGrande_Atual - VolumeU_BarraGrande_COMP])
page_dados.append(['Jacuí',Aflu_Jacui_Atual,Deflu_Jacui_Atual, VolumeU_Jacui_Atual, Maflu_Jacui, Mdeflu_Jacui, VolumeU_Jacui_Atual - VolumeU_Jacui_COMP])
page_dados.append(['Serra da Mesa',Aflu_SerradaMesa_Atual,Deflu_SerradaMesa_Atual,VolumeU_SerradaMesa_Atual, Maflu_SerradaMesa, Mdeflu_SerradaMesa, VolumeU_SerradaMesa_Atual - VolumeU_SerradaMesa_COMP])
page_dados.append(['Tucuruí',Aflu_Tucurui_Atual,Deflu_Tucurui_Atual,VolumeU_Tucurui_Atual, Maflu_Tucurui, Mdeflu_Tucurui, VolumeU_Tucurui_Atual - VolumeU_Tucurui_COMP])
page_dados.append(['Três Marias',Aflu_TrMarias_Atual,Deflu_TrMarias_Atual,VolumeU_TrMarias_Atual, Maflu_TrMarias, Mdeflu_TrMarias, VolumeU_TrMarias_Atual - VolumeU_TrMarias_COMP])
page_dados.append(['Sobradinho',Aflu_Sobradinho_Atual,Deflu_Sobradinho_Atual,VolumeU_Sobradinho_Atual, Maflu_Sobradinho, Mdeflu_Sobradinho, VolumeU_Sobradinho_Atual - VolumeU_Sobradinho_COMP])

print('Dados coletados e salvos com sucesso...')
followSIN.save('FollowSIN.xlsx')
